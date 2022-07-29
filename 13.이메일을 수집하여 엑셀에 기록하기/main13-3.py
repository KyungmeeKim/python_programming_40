from urllib import response
import requests
import re 
from openpyxl import load_workbook
from openpyxl import Workbook
import os

os.chdir(os.path.dirname(os.path.abspath(__file__)))

url = 'https://news.v.daum.net/v/20211129144552297'

headers = {
    'User-Agent' : 'Mozilla/5.0',
    'Content-Type' : 'text/html; charset=utf-8'
}

response = requests.get(url, headers=headers)

results = re.findall(r'[\w\.-]+@[\w\.-]+', response.text)

results = list(set(results))

print(results)

try:  # email.xlsx 파일이 있어 읽어올 수 있다면 파일을 읽습니다.
    wb = load_workbook(r"./email.xlsx", data_only=True)
except:  # email.xlsx 파일이 없다면 새로운 엑셀을 생성합니다.
    wb = Workbook()
    sheet = wb.active

# 이메일을 수집한 수 만큼 반복합니다.
for result in results:
    sheet.append([result])

wb.save(r'./email.xlsx')  # 엑셀 파일을 저장합니다.