from openpyxl import load_workbook  # 엑셀에서 파일을 읽기 위해 openpyxl 라이브러리에서 load_workbook을 불러옵니다.
import os

os.chdir(os.path.dirname(os.path.abspath(__file__)))

load_wb = load_workbook(r"./수료증명단.xlsx")  # 엑셀파일을 읽어옵니다.
load_ws = load_wb.active  # 읽어온 엑셀 파일에서 활성화 된 시트를 불러옵니다.

# 이름, 생년월일, 수료증번호 빈 리스트를 생성합니다.
name_list = []
birthday_list = []
ho_list = []

# 마지막줄까지 for 문을 실행합니다.
for i in range(1, load_ws.max_row + 1):
    name_list.append(load_ws.cell(i, 1).value)  # 1행의 값을 읽어 name_list에 바인딩합니다.
    birthday_list.append(load_ws.cell(i, 2).value)  # 2행의 값을 읽어 birthday_list에 바인딩합니다.
    ho_list.append(load_ws.cell(i, 3).value)  # 3행의 번호를 읽어 ho_list에 바인딩합니다.

print(name_list)
print(birthday_list)
print(ho_list)